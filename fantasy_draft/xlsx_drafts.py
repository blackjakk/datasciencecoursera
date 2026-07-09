"""Parse historical drafts from MONEY_LEAGUE.xlsx using cell fill colors.

Each FF<year> sheet has:
- A draft grid: rows 0-16 = R1-R17, cols B-M = slot 1-12
- A team-color legend at rows 20-31: col A = nickname, col B = color swatch

A cell's BACKGROUND COLOR identifies the manager who actually owned that
pick (post-Yahoo-trade), independent of the original slot. This is the
only reliable source of pick attribution for pre-2024 seasons since the
Sleeper migration lost Yahoo's pick trades.

Public API:
  load_xlsx_drafts(xlsx_path) -> dict[year, list[XlsxPick]]
  XlsxPick = {year, round, slot, player_name, manager_nickname}

To resolve manager_nickname → Sleeper roster_id, use
fantasy_draft.team_identity.manager_for_xlsx_nickname().
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class XlsxPick:
    year: int
    round: int
    slot: int            # original 1-12 draft slot
    player_name: str
    manager_nickname: str  # actual drafter per cell color (post-trade)


def _cell_color(cell) -> str | None:
    f = cell.fill
    if f and f.fgColor and f.fgColor.type == "rgb":
        return f.fgColor.rgb
    return None


_WHITE_LIKE = {"FFFFFFFF", "00000000", "FFFFFF", "000000"}


def _team_color_map(ws, name_col: int = 1, color_col: int = 2,
                    row_range: range = range(20, 32)) -> dict[str, str]:
    """Build {color_hex: manager_nickname} from a range of legend rows.

    If exactly one legend entry has a white-like color (FFFFFFFF), also
    map the "no fill" / transparent code (00000000) to that nickname —
    some older sheets (FF2018) left the slot's draft cells unfilled while
    the legend swatch was explicitly set to white.
    """
    out: dict[str, str] = {}
    white_nick: str | None = None
    for row_idx in row_range:
        nick = ws.cell(row=row_idx + 1, column=name_col).value
        color = _cell_color(ws.cell(row=row_idx + 1, column=color_col))
        if nick and color:
            n = str(nick).strip()
            out[color] = n
            if color in _WHITE_LIKE:
                white_nick = n
    if white_nick and not any(c in out for c in _WHITE_LIKE if c != "FFFFFFFF"):
        for c in _WHITE_LIKE:
            out.setdefault(c, white_nick)
    return out


# Legacy sheet layouts (10-team era, FF2015-FF2018).
# Each entry: (legend_row_range_0idx, legend_name_col, legend_color_col,
#              grid_row_range_0idx, grid_col_range_0idx)
_LEGACY_LAYOUTS = {
    2015: {
        # Names in col L (12), color in col M (13). Rows 1-11 (0-indexed 1..10).
        "legend": (range(1, 11), 12, 13),
        "grid_rows": range(0, 19),
        "grid_cols": range(0, 10),
    },
    2016: {
        # Legend at rows 14-23 (0-idx 13..22), col A name, col B color.
        "legend": (range(13, 23), 1, 2),
        # Round header "1.0" in col D, picks in cols E-N. Rows 1-19 = R1..R19.
        "grid_rows": range(0, 19),
        "grid_cols": range(4, 14),
    },
    2017: {
        "legend": (range(13, 23), 1, 2),
        "grid_rows": range(0, 19),
        "grid_cols": range(4, 14),
    },
    2018: {
        # 2018 used same legend layout as 2016/17 but grid shifted one column.
        "legend": (range(13, 23), 1, 2),
        "grid_rows": range(0, 19),
        "grid_cols": range(5, 15),
    },
}


def _picks_for_legacy_year(ws, year: int) -> list[XlsxPick]:
    layout = _LEGACY_LAYOUTS[year]
    legend_rows, name_col, color_col = layout["legend"]
    color_map = _team_color_map(ws, name_col=name_col, color_col=color_col,
                                 row_range=legend_rows)
    if len(color_map) < 7:  # 10-team era, allow some missing
        return []
    picks: list[XlsxPick] = []
    grid_cols = list(layout["grid_cols"])
    for r in layout["grid_rows"]:
        for slot_idx, c in enumerate(grid_cols):
            cell = ws.cell(row=r + 1, column=c + 1)
            player = cell.value
            if not player:
                continue
            color = _cell_color(cell)
            if color not in color_map:
                continue
            picks.append(XlsxPick(
                year=year, round=r + 1, slot=slot_idx + 1,
                player_name=str(player).strip(),
                manager_nickname=color_map[color],
            ))
    return picks


def load_xlsx_drafts(
    xlsx_path: str | Path = "data/historical/MONEY_LEAGUE.xlsx",
    years: list[int] | None = None,
) -> dict[int, list[XlsxPick]]:
    """Return {year: [XlsxPick]} for all FF<year> sheets we can parse.

    Two layout families:
      - 2019+: 12-team, legend at rows 20-31 cols A-B, grid at rows 1-17
               cols B-M.
      - 2015-2018: 10-team, legend + grid coordinates vary per year
        (see _LEGACY_LAYOUTS).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[int, list[XlsxPick]] = {}
    for sheet in wb.sheetnames:
        if not sheet.startswith("FF"):
            continue
        try:
            year = int(sheet[2:])
        except ValueError:
            continue
        if years and year not in years:
            continue
        ws = wb[sheet]

        if year in _LEGACY_LAYOUTS:
            picks = _picks_for_legacy_year(ws, year)
        else:
            color_map = _team_color_map(ws)
            if len(color_map) < 8:
                continue
            picks = []
            for r in range(17):
                for c in range(12):
                    cell = ws.cell(row=r + 1, column=c + 2)
                    player = cell.value
                    if not player:
                        continue
                    color = _cell_color(cell)
                    if color not in color_map:
                        continue
                    picks.append(XlsxPick(
                        year=year, round=r + 1, slot=c + 1,
                        player_name=str(player).strip(),
                        manager_nickname=color_map[color],
                    ))
        if picks:
            out[year] = picks
    return out
