"""Historical analyses from the 11-year MONEY_LEAGUE xlsx.

Surfaces patterns useful for the 2026 draft:
  - Keeper retention rate by position (QBs kept more often than RBs?)
  - Round value curve (typical career rank of player drafted at round R)
  - Per-team keeper tendencies (which slot each team treats as 'must-keep')
  - 3rd-year cliff: what happens to players AFTER they hit the cap
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .name_aliases import resolve_xlsx_name
from .xlsx_history import load_all_keepers, normalize_name, KeeperRecord


@dataclass
class PlayerCareerLine:
    name: str
    appearances: list[tuple[int, int, int]]  # (year, round, years_kept)


def _detect_layout(ws) -> tuple[int, int]:
    """Find the round-number column and the last round, scanning all columns
    for a 1, 2, 3, ... sequence starting at row 1.

    The sheet layout changed across years:
      - 2019-2025: round number in col 1,    picks in cols 2..13.
      - 2018:      round number in col 5,    picks in cols 6..15.
      - 2016-2017: round number in col 4,    picks in cols 5..14.
      - 2015:      NO round column at all;   picks in cols 1..10, rows ARE
                   rounds. Falls back to round_col=0 (no skip), max=19.

    openpyxl returns numbers as floats (1.0, 2.0, ...). Returns
    (round_col, last_round). round_col=0 means "no round col; don't skip
    any columns on the left."
    """
    for col_idx in range(1, ws.max_column + 1):
        contiguous = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, (int, float)) and float(v) == contiguous + 1:
                contiguous += 1
            else:
                break
        if contiguous >= 10:  # need at least a 10-round draft to count
            return col_idx, contiguous
    # No round column (older 2015 layout). Use max practical draft size as cap.
    return 0, 19


def load_full_grid(path: str | Path) -> dict[int, dict[tuple[int, int], dict]]:
    """{year: {(round, col): {name, color, comment}}} for every cell with a player.

    Only cells inside the actual draft grid are read (rows 1..last_round,
    cols immediately right of the round column). Cells below the draft
    (legend like "Color"/"Hotel"/"Dues", commissioner notes, free-form
    comments) and cells to the left (team-name roster legend in older
    sheets) are excluded.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[int, dict[tuple[int, int], dict]] = {}
    for sheet in wb.sheetnames:
        m = re.match(r'FF(\d{4})$', sheet)
        if not m:
            continue
        year = int(m.group(1))
        ws = wb[sheet]
        round_col, max_round = _detect_layout(ws)
        grid: dict[tuple[int, int], dict] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not str(cell.value).strip():
                    continue
                if cell.column <= round_col:  # round column + everything left of it
                    continue
                if cell.row > max_round:  # below the draft grid -- legend/notes
                    continue
                if not isinstance(cell.value, str):
                    continue
                canonical = resolve_xlsx_name(cell.value)
                if canonical is None:
                    continue  # owner name / admin label
                color = None
                try:
                    color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                except Exception:
                    pass
                comment = cell.comment.text if cell.comment else ''
                yr_n = 0
                if 'keeper' in comment.lower():
                    cm = re.search(r'(\d+)', comment)
                    yr_n = int(cm.group(1)) if cm else 1
                grid[(cell.row, cell.column)] = {
                    'name': canonical,
                    'raw_name': cell.value.strip(),
                    'color': color,
                    'years_kept': yr_n,
                    'comment': comment.strip(),
                }
        out[year] = grid
    return out


def keeper_retention_by_position(path: str | Path,
                                  sleeper_picks: list[dict] | None = None) -> dict[str, dict]:
    """For each position, what % of 1st-year keepers become 2nd-year? 2nd → 3rd?

    Builds career timelines from player-name matching across years, then counts
    consecutive keeper streaks.

    Uses Sleeper picks (if provided) for position metadata. Otherwise, scan
    historical names against a small heuristic isn't reliable — pass picks.
    """
    by_year = load_all_keepers(path)
    # Build {norm_name: [(year, years_kept)]} chronologically.
    timeline: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for year, kps in by_year.items():
        for k in kps:
            timeline[normalize_name(k.player_name)].append((year, k.years_kept))

    # Position lookup from Sleeper picks (current names).
    pos_lookup: dict[str, str] = {}
    if sleeper_picks:
        for p in sleeper_picks:
            name = f"{p['metadata']['first_name']} {p['metadata']['last_name']}".strip()
            pos_lookup[normalize_name(name)] = p['metadata']['position']

    # Count transitions: yr1→yr2, yr2→yr3, yr3→dropped.
    transitions: dict[str, Counter] = defaultdict(Counter)
    for name, hits in timeline.items():
        pos = pos_lookup.get(name, '?')
        hits.sort()
        # Walk chronologically; for each year, look at next year to see if continued.
        for i, (year, yr_n) in enumerate(hits):
            if yr_n == 3:
                transitions[pos]['hit_cap'] += 1
                continue
            # Did they appear again next year as keeper?
            next_year = year + 1
            next_entry = next(((y, n) for y, n in hits if y == next_year), None)
            if next_entry and next_entry[1] == yr_n + 1:
                transitions[pos][f'yr{yr_n}_continued'] += 1
            else:
                transitions[pos][f'yr{yr_n}_dropped'] += 1

    # Compute retention rates.
    result: dict[str, dict] = {}
    for pos in sorted(set(list(transitions.keys()) + ['QB', 'RB', 'WR', 'TE'])):
        t = transitions[pos]
        y1_total = t['yr1_continued'] + t['yr1_dropped']
        y2_total = t['yr2_continued'] + t['yr2_dropped']
        result[pos] = {
            'yr1_count': y1_total,
            'yr1_to_yr2_pct': (100 * t['yr1_continued'] / y1_total) if y1_total else None,
            'yr2_count': y2_total,
            'yr2_to_yr3_pct': (100 * t['yr2_continued'] / y2_total) if y2_total else None,
            'hit_cap_count': t['hit_cap'],
        }
    return result


def team_keeper_tendencies(path: str | Path) -> dict[int, dict]:
    """For each team-column, count keepers per year + average keepers per
    position. Helps identify "QB hoarder" or "always keeps 4" teams."""
    grid = load_full_grid(path)
    # team_col -> {year: count_keepers, position_keeper_counts}
    by_col: dict[int, dict] = defaultdict(lambda: {
        'keepers_per_year': defaultdict(int),
        'positions_kept': Counter(),
        'avg_keeper_round': [],
    })
    for year, cells in grid.items():
        for (r, c), data in cells.items():
            if data['years_kept'] > 0:
                by_col[c]['keepers_per_year'][year] += 1
                by_col[c]['avg_keeper_round'].append(r)
    # Compute summaries.
    out: dict[int, dict] = {}
    for col, info in by_col.items():
        per_yr = info['keepers_per_year']
        avg_round = (sum(info['avg_keeper_round']) / len(info['avg_keeper_round'])
                     if info['avg_keeper_round'] else 0)
        out[col] = {
            'total_keepers_all_years': sum(per_yr.values()),
            'avg_keepers_per_year': sum(per_yr.values()) / max(1, len(per_yr)),
            'avg_keeper_round': round(avg_round, 1),
            'most_recent_5yrs': {y: per_yr.get(y, 0) for y in sorted(per_yr)[-5:]},
        }
    return out


def post_cap_dropoff(path: str | Path,
                     sleeper_picks: list[dict] | None = None) -> list[dict]:
    """For each player who hit the 3-year cap, do they reappear next year (as a
    fresh draft pick) at an earlier round (still good), later round (declining),
    or not at all (cut/retired)?"""
    by_year = load_all_keepers(path)
    grid = load_full_grid(path)
    pos_lookup: dict[str, str] = {}
    if sleeper_picks:
        for p in sleeper_picks:
            name = f"{p['metadata']['first_name']} {p['metadata']['last_name']}".strip()
            pos_lookup[normalize_name(name)] = p['metadata']['position']

    out: list[dict] = []
    for year, kps in by_year.items():
        for k in kps:
            if k.years_kept != 3:
                continue
            next_year = year + 1
            # Find player in next year's grid.
            next_grid = grid.get(next_year, {})
            next_appearance = None
            for (r, c), data in next_grid.items():
                if normalize_name(data['name']) == normalize_name(k.player_name):
                    next_appearance = (r, c)
                    break
            # "Earlier" = redrafted in a strictly earlier round than where they
            # were kept (held or grew their value). The old threshold
            # `next_year_round < k.round_num - 2` had two problems: (1) it
            # confused round_penalty with delta-detection so an R3 keeper
            # required next_year_round < 1 (impossible), silently routing
            # every early-round capped player to "redrafted_later"; (2) it
            # treated "same round" as "later" rather than "same value".
            if not next_appearance:
                fate = 'undrafted_next_year'
            elif next_appearance[0] < k.round_num:
                fate = 'redrafted_earlier'
            elif next_appearance[0] == k.round_num:
                fate = 'redrafted_same_round'
            else:
                fate = 'redrafted_later'
            out.append({
                'capped_year': year,
                'player': k.player_name,
                'position': pos_lookup.get(normalize_name(k.player_name), '?'),
                'kept_at_round': k.round_num,
                'next_year_round': next_appearance[0] if next_appearance else None,
                'next_year_col': next_appearance[1] if next_appearance else None,
                'fate': fate,
            })
    return out
