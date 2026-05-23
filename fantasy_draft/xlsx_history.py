"""Parse the MONEY_LEAGUE.xlsx historical-draft spreadsheet.

Layout:
  - One tab per year (FF2025, FF2024, ..., FF2015).
  - Column 1: round number (1..17).
  - Columns 2..13: 12 teams, in draft-slot order.
  - Cell value: player name drafted at that slot/round.
  - Cell fill color: marks which team currently owns the pick (trades show as
    off-color cells).
  - Cell comment: "Nth year keeper" tags keep status + how many consecutive
    seasons. THIS is the authoritative keeper record — more complete than
    Sleeper's is_keeper flag.

Use load_keepers_for_year() to extract per-year keeper records (player,
years_kept, round). Use load_full_draft() if you need every pick.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .name_aliases import resolve_xlsx_name


@dataclass
class KeeperRecord:
    year: int
    round_num: int
    column: int            # spreadsheet column 2..13 (= draft_slot + 1)
    player_name: str       # canonical name (post-alias resolution)
    raw_xlsx_name: str     # original cell value before alias resolution
    years_kept: int        # 1, 2, or 3 (3 = hits the cap, can't keep next year)
    raw_note: str


def load_keepers_for_year(path: str | Path, year: int) -> list[KeeperRecord]:
    """Extract all keeper-tagged players from a single tab.

    Player names are passed through name_aliases.resolve_xlsx_name so they
    join cleanly to the Sleeper catalog. The original xlsx string is kept
    in raw_xlsx_name for debugging.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = f"FF{year}"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No tab '{sheet_name}' in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    out: list[KeeperRecord] = []
    for row in ws.iter_rows():
        for cell in row:
            if not cell.comment or 'keeper' not in cell.comment.text.lower():
                continue
            if cell.value is None:
                continue
            raw = str(cell.value).strip()
            canonical = resolve_xlsx_name(raw)
            if canonical is None:
                # Owner-name or admin string with a stray "keeper" comment.
                continue
            m = re.search(r'(\d+)', cell.comment.text)
            yr_n = int(m.group(1)) if m else 1
            out.append(KeeperRecord(
                year=year,
                round_num=cell.row,
                column=cell.column,
                player_name=canonical,
                raw_xlsx_name=raw,
                years_kept=yr_n,
                raw_note=cell.comment.text.strip(),
            ))
    return out


def load_all_keepers(path: str | Path) -> dict[int, list[KeeperRecord]]:
    """{year: [KeeperRecord]} for every FFYYYY tab present."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[int, list[KeeperRecord]] = {}
    for name in wb.sheetnames:
        m = re.match(r'FF(\d{4})$', name)
        if not m:
            continue
        year = int(m.group(1))
        out[year] = load_keepers_for_year(path, year)
    return out


def normalize_name(name: str) -> str:
    return (name.lower().replace(".", "").replace("'", "")
            .replace("-", " ").replace(" jr", "").replace(" sr", "")
            .replace(" iii", "").replace(" ii", "").strip())
